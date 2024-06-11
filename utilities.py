from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO
import pandas as pd

def extraire_titres_numerotes(doc_file):
    from docx import Document
    doc = Document(doc_file)
    data = []
    num_titre = [0] * 10

    for para in doc.paragraphs:
        if para.style.name.startswith('Heading'):
            niveau = int(para.style.name.replace('Heading ', ''))
            num_titre[niveau - 1] += 1
            for i in range(niveau, 10):
                num_titre[i] = 0
            numero_complet = '.'.join(str(num) for num in num_titre[:niveau] if num > 0)
            data.append((numero_complet, para.text))
    return pd.DataFrame(data, columns=['N°', 'DESIGNATION DES OUVRAGES'])

def add_data_to_existing_excel(df, cell_A2_content, cell_C2_content, feuille,feuille_, existing_file):
    # Charger le fichier Excel existant depuis un objet BytesIO
    wb = load_workbook(existing_file)

    ws = wb[feuille]
    # Styles
    title_font = Font(name='Arial', size=9, bold=True)
    content_font = Font(name='Arial', size=9)
    center_aligned_text = Alignment(horizontal='center', vertical='center')
    left_aligned_text = Alignment(horizontal='left', vertical='center')

    # Mettre à jour le contenu des cellules A2 et C2
    ws['A2'] = cell_A2_content
    ws['A2'].font = title_font
    ws['A2'].alignment = center_aligned_text

    ws['C2'] = 'Lot '+cell_C2_content
    ws['C2'].font = title_font
    ws['C2'].alignment = center_aligned_text

    # Assurez-vous que le DataFrame a au moins autant de lignes que nécessaire
    start_row = 4
    end_row = 63

    if len(df) > (end_row - start_row + 1):
        raise ValueError("Le DataFrame contient plus de lignes que la plage cible.")

    # Ajouter les données du DataFrame dans la plage spécifiée
    for i, (index, row) in enumerate(df.iterrows(), start=start_row):
        if i > end_row:
            break
        ws[f'A{i}'] = index
        ws[f'A{i}'].font = content_font
        ws[f'A{i}'].alignment = left_aligned_text

        ws[f'B{i}'] = row['DESIGNATION DES OUVRAGES']
        ws[f'B{i}'].font = content_font
        ws[f'B{i}'].alignment = left_aligned_text
    ws.title = feuille_
    # Sauvegarder le fichier mis à jour dans un BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
