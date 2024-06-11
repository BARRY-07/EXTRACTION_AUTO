from docx import Document
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, Border, Side, Font, NamedStyle
from io import BytesIO

def extraire_titres_numerotes(doc_file):
    doc = Document(doc_file)
    data = []
    num_titre = [0] * 10

    for para in doc.paragraphs:
        if para.style.name.startswith('Heading'):
            niveau = int(para.style.name.replace('Heading ', ''))
            num_titre[niveau - 1] += 1
            for i in range(niveau, 10):
                num_titre[i] = 0
            numero_complet = '.'.join(str(num) for num in num_titre[:niveau] if num > 0)
            data.append((numero_complet, para.text))
    return pd.DataFrame(data, columns=['N°', 'DESIGNATION DES OUVRAGES'])


def convert_df_to_excel(df, cell_A2_content, cell_C2_content,feuille):
    title_font = Font(name='Arial', size=9, bold=True)
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = feuille

    # Titre et sous-titre
    ws.append(['DECOMPOSITION DU PRIX GLOBAL ET FORFAITAIRE'])
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')

    ws.append([cell_A2_content])
    ws['A2'].font = title_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A2:B2')

    ws['C2'].value = 'Lot '+cell_C2_content
    ws['C2'].font = title_font
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Appliquer le remplissage de couleur à la cellule C2
    ws['C2'].fill = yellow_fill
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C2:F2')

    # Ajout des en-têtes de colonnes
    ws.append(['N°', 'DESIGNATION DES OUVRAGES', 'U', 'QT', 'P.U.', 'PRIX TOTAUX'])
    ws['A3'].font=title_font
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B3'].font=title_font
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C3'].font=title_font
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D3'].font=title_font
    ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E3'].font=title_font
    ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F3'].font=title_font
    ws['F3'].alignment = Alignment(horizontal='center', vertical='center')

    # Dimensions des lignes et colonnes
    ws.row_dimensions[1].height = 19.5
    ws.row_dimensions[2].height = 48.75
    ws.row_dimensions[3].height = 24

    ws.column_dimensions['A'].width = 7.43
    ws.column_dimensions['B'].width = 47
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 9.29
    ws.column_dimensions['E'].width = 10.57
    ws.column_dimensions['F'].width = 14.14





    # Ajouter les données du DataFrame
    for index, row in df.iterrows():
        ws.append([row['N°'], row['DESIGNATION DES OUVRAGES'], '', '', '', ''])

    ws['B39'].value = 'TOTAL GENERAL HT'
    ws['B39'].alignment = Alignment(horizontal='right', vertical='center')
    ws['B39'].font = title_font

    ws['B41'].value = 'TVA 20 %'
    ws['B41'].alignment = Alignment(horizontal='right', vertical='center')
    ws['B41'].font = title_font

    ws['B43'].value = 'TOTAL GENERAL TTC'
    ws['B43'].font = title_font
    ws['B43'].alignment = Alignment(horizontal='right', vertical='center')  


    thin_side = Side(style='thin')
    
    
    def apply_outer_borders(range_str):
        rows = ws[range_str]
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                border = Border(
                    left=thin_side if j == 0 else None,
                    right=thin_side if j == len(row) - 1 else None,
                    top=thin_side if i == 0 else None,
                    bottom=thin_side if i == len(rows) - 1 else None
                )
                cell.border = border

    apply_outer_borders('A1:F45')
    apply_outer_borders('A4:A45')
    apply_outer_borders('B4:B37') 
    apply_outer_borders('C4:C37')  
    apply_outer_borders('D4:D37')  
    apply_outer_borders('E4:E37')  
    apply_outer_borders('F4:F37')

    apply_outer_borders('B38:F44')


    apply_outer_borders('C38:C44')
    apply_outer_borders('D38:D44')
    apply_outer_borders('E38:E44')
    apply_outer_borders('F38:F44')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border

    
    wb.save(output)
    return output.getvalue()
